import streamlit as st
import sqlite3
import re
import pandas as pd
import os
import zipfile
import io
import shutil
from datetime import datetime
import pypdf

# BIBLIOTECAS PARA RELATÓRIOS
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# CONFIGURAÇÃO DE CAMINHO AUTOMÁTICO
DIRETORIO_ATUAL = os.path.dirname(os.path.abspath(__file__))
CAMINHO_BANCO = os.path.join(DIRETORIO_ATUAL, "diarias_sistema.db")
CAMINHO_BRASAO = os.path.join(DIRETORIO_ATUAL, "brasao_pr.png")

# Pasta de empenhos
PASTA_EMPENHOS = os.path.join(DIRETORIO_ATUAL, "arquivos_empenhos")
if not os.path.exists(PASTA_EMPENHOS):
    os.makedirs(PASTA_EMPENHOS)

# Subpasta para onde são movidos os PDFs de empenhos com saldo zerado (aguardando exclusão futura)
PASTA_EMPENHOS_ZERADOS = os.path.join(PASTA_EMPENHOS, "EMPENHOS ZERADOS")
if not os.path.exists(PASTA_EMPENHOS_ZERADOS):
    os.makedirs(PASTA_EMPENHOS_ZERADOS)

# 1. CONFIGURAÇÃO DO BANCO DE DADOS
def conectar_banco():
    conn = sqlite3.connect(CAMINHO_BANCO)
    cursor = conn.cursor()
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS servidores (
            id INTEGER PRIMARY KEY,
            nome_completo TEXT NOT NULL,
            cpf TEXT UNIQUE NOT NULL,
            unidade_lotacao TEXT NOT NULL,
            banco_codigo TEXT NOT NULL,
            agencia TEXT NOT NULL,
            conta_corrente TEXT NOT NULL,
            status TEXT DEFAULT 'Ativo'
        )
    """)
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS unidades (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome_unidade TEXT UNIQUE NOT NULL
        )
    """)
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS bancos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT UNIQUE NOT NULL,
            nome_banco TEXT NOT NULL
        )
    """)
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS empenhos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            servidor_nome TEXT,
            numero_empenho TEXT UNIQUE NOT NULL,
            valor_total REAL NOT NULL,
            valor_disponivel REAL NOT NULL,
            data_empenho TEXT,
            caminho_pdf TEXT
        )
    """)
    
    cursor.execute("PRAGMA table_info(empenhos)")
    colunas = [coluna[1] for coluna in cursor.fetchall()]
    
    if "servidor_nome" not in colunas:
        cursor.execute("ALTER TABLE empenhos ADD COLUMN servidor_nome TEXT")
    if "data_empenho" not in colunas:
        cursor.execute("ALTER TABLE empenhos ADD COLUMN data_empenho TEXT")
    if "caminho_pdf" not in colunas:
        cursor.execute("ALTER TABLE empenhos ADD COLUMN caminho_pdf TEXT")
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS diarias (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            servidor_id INTEGER NOT NULL,
            empenho_id INTEGER NOT NULL,
            data_diaria TEXT NOT NULL,
            jornada TEXT NOT NULL,
            valor_diaria REAL NOT NULL,
            FOREIGN KEY(servidor_id) REFERENCES servidores(id),
            FOREIGN KEY(empenho_id) REFERENCES empenhos(id)
        )
    """)
    
    cursor.execute("SELECT COUNT(*) FROM bancos")
    if cursor.fetchone()[0] == 0:
        bancos_iniciais = [
            ("001", "BANCO DO BRASIL S.A."),
            ("104", "CAIXA ECONOMICA CEF"),
            ("341", "ITAU UNIBANCO S.A."),
            ("237", "BANCO BRADESCO S.A."),
            ("033", "BANCO SANTANDER (BRASIL) S.A.")
        ]
        cursor.executemany("INSERT INTO bancos (codigo, nome_banco) VALUES (?, ?)", bancos_iniciais)
        
    cursor.execute("SELECT COUNT(*) FROM unidades")
    if cursor.fetchone()[0] == 0:
        unidades_iniciais = [("Sede DEPEN"), ("CCCE"), ("CPA"), ("Complexo Piraquara")]
        cursor.executemany("INSERT INTO unidades (nome_unidade) VALUES (?)", [(u,) for u in unidades_iniciais])
    
    conn.commit()
    return conn, cursor

# Move o PDF do empenho para a pasta "EMPENHOS ZERADOS" quando o saldo chega a zero,
# e traz de volta para a pasta principal caso o saldo volte a ficar positivo (ex: edição ou exclusão de diária).
def sincronizar_pasta_empenho(cursor, id_empenho):
    try:
        cursor.execute("SELECT valor_disponivel, caminho_pdf FROM empenhos WHERE id = ?", (id_empenho,))
        resultado = cursor.fetchone()
        if not resultado:
            return
        saldo_atual, caminho_pdf_atual = resultado
        if not caminho_pdf_atual or not os.path.exists(caminho_pdf_atual):
            return
        
        pasta_atual = os.path.dirname(os.path.abspath(caminho_pdf_atual))
        nome_arquivo = os.path.basename(caminho_pdf_atual)
        esta_zerado = saldo_atual is not None and round(float(saldo_atual), 2) <= 0
        
        if esta_zerado and pasta_atual != os.path.abspath(PASTA_EMPENHOS_ZERADOS):
            novo_caminho = os.path.join(PASTA_EMPENHOS_ZERADOS, nome_arquivo)
            shutil.move(caminho_pdf_atual, novo_caminho)
            cursor.execute("UPDATE empenhos SET caminho_pdf = ? WHERE id = ?", (novo_caminho, id_empenho))
        elif not esta_zerado and pasta_atual == os.path.abspath(PASTA_EMPENHOS_ZERADOS):
            novo_caminho = os.path.join(PASTA_EMPENHOS, nome_arquivo)
            shutil.move(caminho_pdf_atual, novo_caminho)
            cursor.execute("UPDATE empenhos SET caminho_pdf = ? WHERE id = ?", (novo_caminho, id_empenho))
    except Exception:
        # Falha ao mover o arquivo não deve interromper o lançamento/edição da diária
        pass

# Extração de PDF
def extrair_dados_pdf_pr(arquivo_pdf):
    try:
        leitor = pypdf.PdfReader(arquivo_pdf)
        texto = ""
        for pagina in leitor.pages:
            texto += pagina.extract_text() + "\n"
            
        dados = {"empenho": "", "data": datetime.now().date(), "servidor": "", "valor": 0.0}
        
        match_emp = re.search(r'(\d{4}NE\d{6})', texto)
        if match_emp:
            dados["empenho"] = match_emp.group(1).strip().upper()
            
        match_dt = re.search(r'\d{4}NE\d{6}\s+(\d{2}/\d{2}/\d{2,4})', texto)
        if match_dt:
            dt_str = match_dt.group(1)
            try:
                if len(dt_str.split('/')[-1]) == 2:
                    dados["data"] = datetime.strptime(dt_str, "%d/%m/%y").date()
                else:
                    dados["data"] = datetime.strptime(dt_str, "%d/%m/%Y").date()
            except:
                pass
                
        match_cred = re.search(r'Credor\s+\d+\s+-\s+([A-Z\sÀ-Ú]+)', texto)
        if match_cred:
            dados["servidor"] = match_cred.group(1).strip().upper()
            
        match_vl = re.search(r'Valor\s+([\d\.,]+)', texto)
        if match_vl:
            vl_str = match_vl.group(1).replace('.', '').replace(',', '.')
            dados["valor"] = float(vl_str)
            
        return dados
    except Exception as e:
        st.error(f"Erro ao processar a leitura do PDF: {e}")
        return None

# Funções auxiliares de listagem
def listar_bancos_cadastrados():
    conn, cursor = conectar_banco()
    cursor.execute("SELECT codigo, nome_banco FROM bancos ORDER BY codigo ASC")
    bancos = [f"{linha[0]} - {linha[1]}" for linha in cursor.fetchall()]
    conn.close()
    return bancos

def buscar_unidades_dict():
    conn, cursor = conectar_banco()
    cursor.execute("SELECT id, nome_unidade FROM unidades ORDER BY nome_unidade ASC")
    unidades = {linha[1]: linha[0] for linha in cursor.fetchall()}
    conn.close()
    return unidades

def listar_servidores_selecao():
    conn, cursor = conectar_banco()
    cursor.execute("SELECT id, nome_completo FROM servidores WHERE status = 'Ativo' ORDER BY nome_completo ASC")
    servidores = [{ "id": id_linha, "label": nome_linha } for id_linha, nome_linha in cursor.fetchall()]
    conn.close()
    return servidores

def listar_todos_empenhos():
    conn, cursor = conectar_banco()
    cursor.execute("SELECT id, numero_empenho, servidor_nome FROM empenhos ORDER BY numero_empenho ASC")
    linhas = cursor.fetchall()
    conn.close()
    return linhas

def listar_empenhos_com_saldo():
    conn, cursor = conectar_banco()
    cursor.execute("SELECT id, numero_empenho, valor_disponivel, servidor_nome FROM empenhos WHERE valor_disponivel > 0 ORDER BY numero_empenho ASC")
    linhas = cursor.fetchall()
    conn.close()
    
    empenhos = []
    for Proxy_linha in linhas:
        id_emp = Proxy_linha[0]
        num_emp = Proxy_linha[1]
        saldo = Proxy_linha[2]
        servidor = Proxy_linha[3] if Proxy_linha[3] else "Sem vínculo"
        saldo_formatado = f"R$ {saldo:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        empenhos.append({
            "id": id_emp,
            "label": f"NE: {num_emp} | {servidor} | Saldo: {saldo_formatado}"
        })
    return empenhos

# Formatação numérica no padrão brasileiro (sem "R$")
def formatar_valor_br(valor):
    return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

# Formatação com "R$" (usada na planilha DEAEV REALIZADA)
def formatar_valor_br_moeda(valor):
    return f"R$ {formatar_valor_br(valor)}"

# Formata quantidade de DEAEVs: inteiro sem casas decimais, fracionado sem zero à direita (ex: 2,5 em vez de 2,50)
def formatar_qtd_deaev(valor):
    if valor == int(valor):
        return str(int(valor))
    texto = f"{valor:.2f}".rstrip("0").rstrip(".")
    return texto.replace(".", ",")

# Nomes dos meses por extenso
MESES_EXTENSO = {
    "01": "JANEIRO", "02": "FEVEREIRO", "03": "MARÇO", "04": "ABRIL",
    "05": "MAIO", "06": "JUNHO", "07": "JULHO", "08": "AGOSTO",
    "09": "SETEMBRO", "10": "OUTUBRO", "11": "NOVEMBRO", "12": "DEZEMBRO"
}

# CONTROLE DE MENU ATIVO
if "menu_ativo" not in st.session_state:
    st.session_state.menu_ativo = "inicio"

# Estados da leitura de PDF
if "pdf_num_empenho" not in st.session_state:
    st.session_state.pdf_num_empenho = ""
if "pdf_data_empenho" not in st.session_state:
    st.session_state.pdf_data_empenho = datetime.now().date()
if "pdf_nome_servidor" not in st.session_state:
    st.session_state.pdf_nome_servidor = ""
if "pdf_valor_empenho" not in st.session_state:
    st.session_state.pdf_valor_empenho = 0.0

st.set_page_config(page_title="Sistema de Diárias - Polícia Penal", layout="wide")

# CSS customizado para os botões do menu
st.markdown("""
    <style>
    div[data-testid="stVerticalBlockBorderWrapper"] div[data-testid="stButton"] button {
        width: 210px !important;
        min-width: 210px !important;
        max-width: 210px !important;
        height: 50px !important;
        text-align: left !important;
        padding-left: 18px !important;
        white-space: nowrap !important;
        display: block !important;
    }
    div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] button {
        width: 160px !important;
        min-width: 160px !important;
        max-width: 160px !important;
    }
    </style>
""", unsafe_allow_html=True)

conectar_banco()

# Topo do Sistema
col_logo, col_titulo = st.columns([1, 11])
with col_logo:
    if os.path.exists(CAMINHO_BRASAO):
        st.image(CAMINHO_BRASAO, width=80)
    else:
        st.subheader("🏛️")
with col_titulo:
    st.markdown("<h1 style='margin-top: -5px; padding-bottom: 0px;'>Sistema de Diárias Extraordinárias</h1>", unsafe_allow_html=True)
    st.markdown("<h4 style='color: #4F8BF9; margin-top: -15px; font-weight: bold; letter-spacing: 1px;'>POLÍCIA PENAL — ESTADO DO PARANÁ</h4>", unsafe_allow_html=True)
    
st.markdown("---")

col_menu_botoes, col_conteudo_dinamico = st.columns([2.5, 9.5])

# MENU LATERAL ESQUERDO
with col_menu_botoes:
    st.write("### Módulos")
    
    tipo_btn_servidores = "primary" if st.session_state.menu_ativo == "servidores" else "secondary"
    if st.button("👥 SERVIDORES", type=tipo_btn_servidores, key="btn_menu_servidores"):
        st.session_state.menu_ativo = "servidores"
        st.rerun()
        
    tipo_btn_empenhos = "primary" if st.session_state.menu_ativo == "empenhos" else "secondary"
    if st.button("📝 EMPENHOS", type=tipo_btn_empenhos, key="btn_menu_empenhos"):
        st.session_state.menu_ativo = "empenhos"
        st.rerun()
        
    tipo_btn_deaev = "primary" if st.session_state.menu_ativo == "deaev" else "secondary"
    if st.button("📋 DEAEV", type=tipo_btn_deaev, key="btn_menu_deaev"):
        st.session_state.menu_ativo = "deaev"
        st.rerun()

    tipo_btn_relatorios = "primary" if st.session_state.menu_ativo == "relatorios" else "secondary"
    if st.button("📊 RELATÓRIOS", type=tipo_btn_relatorios, key="btn_menu_relatorios"):
        st.session_state.menu_ativo = "relatorios"
        st.rerun()
        
    st.markdown("---")
    tipo_btn_sistema = "primary" if st.session_state.menu_ativo == "sistema" else "secondary"
    if st.button("⚙️ SISTEMA", type=tipo_btn_sistema, key="btn_menu_sistema"):
        st.session_state.menu_ativo = "sistema"
        st.rerun()

# CONTEÚDO DINÂMICO CENTRAL
with col_conteudo_dinamico:
    dict_unidades = buscar_unidades_dict()
    lista_unidades = list(dict_unidades.keys())
    lista_bancos = listar_bancos_cadastrados()
    lista_servidores = listar_servidores_selecao()
    
    if st.session_state.menu_ativo == "inicio":
        st.info("💡 Selecione um módulo ao lado esquerdo para exibir as opções e ferramentas na tela.")
        
    # =====================================================================
    # MÓDULO SERVIDORES
    # =====================================================================
    elif st.session_state.menu_ativo == "servidores":
        st.markdown("<h2>👥 Gerenciamento de Servidores</h2>", unsafe_allow_html=True)
        
        conn, cursor = conectar_banco()
        cursor.execute("SELECT id, cpf, nome_completo, unidade_lotacao, banco_codigo, agencia, conta_corrente, status FROM servidores ORDER BY nome_completo ASC")
        servidores_gravados = cursor.fetchall()
        conn.close()
        
        col_cadastro, col_consulta = st.columns([1.1, 1.1])
        servidor_selecionado_df = None
        
        with col_consulta:
            st.markdown("#### 🔍 Consulta de Servidores")
            st.caption("Clique no círculo ao lado esquerdo do CPF para editar ou excluir os dados do servidor.")
            
            if servidores_gravados:
                df_completo = pd.DataFrame(servidores_gravados, columns=["ID", "CPF", "Nome Completo", "Lotação", "Banco", "Agência", "Conta Corrente", "Status"])
                df_exibicao = df_completo[["CPF", "Nome Completo"]]
                
                tabela_interativa = st.dataframe(
                    df_exibicao, 
                    use_container_width=True, 
                    hide_index=True, 
                    height=400,
                    on_select="rerun",
                    selection_mode="single-row"
                )
                
                indices_selecionados = tabela_interativa.get("selection", {}).get("rows", [])
                if indices_selecionados:
                    index_linha = indices_selecionados[0]
                    servidor_selecionado_df = df_completo.iloc[index_linha]
            else:
                st.info("Nenhum servidor cadastrado até o momento.")
                
        with col_cadastro:
            if servidor_selecionado_df is not None:
                st.markdown("#### ✏️ Alterar / Excluir Servidor")
                id_atual = int(servidor_selecionado_df["ID"])
                nome_padrao = str(servidor_selecionado_df["Nome Completo"])
                cpf_padrao = str(servidor_selecionado_df["CPF"])
                
                lot_ori = str(servidor_selecionado_df["Lotação"])
                idx_lotacao = lista_unidades.index(lot_ori) if lot_ori in lista_unidades else 0
                
                bnc_ori = str(servidor_selecionado_df["Banco"])
                idx_banco = 0
                for idx, bnc in enumerate(lista_bancos):
                    if bnc.startswith(bnc_ori):
                        idx_banco = idx
                        break
                        
                agencia_padrao = str(servidor_selecionado_df["Agência"])
                conta_padrao = str(servidor_selecionado_df["Conta Corrente"])
                status_padrao = str(servidor_selecionado_df["Status"])
            else:
                st.markdown("#### ➕ Cadastrar Novo Servidor")
                id_atual = None
                nome_padrao = ""
                cpf_padrao = ""
                idx_lotacao = 0
                idx_banco = 0
                agencia_padrao = ""
                conta_padrao = ""
                status_padrao = "Ativo"

            nome_serv = st.text_input("Nome Completo", value=nome_padrao)
            cpf_serv = st.text_input("CPF", value=cpf_padrao)
            unidade_serv = st.selectbox("Unidade de Lotação", options=lista_unidades if lista_unidades else ["Padrão"], index=idx_lotacao)
            banco_serv = st.selectbox("Banco para depósito", options=lista_bancos, index=idx_banco)
            agencia_serv = st.text_input("Agência", value=agencia_padrao)
            conta_serv = st.text_input("Conta Corrente (Até 13 dígitos)", value=conta_padrao, max_chars=13)
            status_serv = st.selectbox("Status", options=["Ativo", "Inativo"], index=0 if status_padrao == "Ativo" else 1)
            
            if id_atual is None:
                if st.button("➕ Salvar Novo Servidor", type="primary", key="btn_salvar_novo_servidor"):
                    if not nome_serv or not cpf_serv or not conta_serv:
                        st.error("Nome, CPF e Conta Corrente são obrigatórios.")
                    else:
                        conta_formatada = conta_serv.strip().zfill(13)
                        conn, cursor = conectar_banco()
                        try:
                            cursor.execute("""
                                INSERT INTO servidores (nome_completo, cpf, unidade_lotacao, banco_codigo, agencia, conta_corrente, status)
                                VALUES (?, ?, ?, ?, ?, ?, ?)
                            """, (nome_serv.upper(), cpf_serv.strip(), unidade_serv, banco_serv.split(" - ")[0], agencia_serv.strip(), conta_formatada, status_serv))
                            conn.commit()
                            st.success(f"Servidor cadastrado com sucesso! Conta salva: {conta_formatada}")
                        except sqlite3.IntegrityError:
                            st.error("CPF já existente no banco de dados.")
                        finally:
                            conn.close()
                            st.rerun()
            else:
                c_atualizar, c_deletar = st.columns(2)
                with c_atualizar:
                    if st.button("💾 Atualizar Dados", type="primary", key="btn_atualizar_servidor"):
                        if not nome_serv or not cpf_serv or not conta_serv:
                            st.error("Nome, CPF e Conta Corrente são obrigatórios.")
                        else:
                            conta_formatada = conta_serv.strip().zfill(13)
                            conn, cursor = conectar_banco()
                            try:
                                nome_novo_formatado = nome_serv.upper()
                                cursor.execute("""
                                    UPDATE servidores 
                                    SET nome_completo = ?, cpf = ?, unidade_lotacao = ?, banco_codigo = ?, agencia = ?, conta_corrente = ?, status = ?
                                    WHERE id = ?
                                """, (nome_novo_formatado, cpf_serv.strip(), unidade_serv, banco_serv.split(" - ")[0], agencia_serv.strip(), conta_formatada, status_serv, id_atual))

                                # Propaga a correção do nome para os empenhos já lançados em nome deste servidor.
                                # (As diárias/DEAEV já são vinculadas por servidor_id via JOIN e se atualizam automaticamente,
                                # mas os empenhos guardam o nome do servidor como texto e precisam ser sincronizados aqui.)
                                if nome_novo_formatado != nome_padrao:
                                    cursor.execute(
                                        "UPDATE empenhos SET servidor_nome = ? WHERE servidor_nome = ?",
                                        (nome_novo_formatado, nome_padrao)
                                    )

                                conn.commit()
                                st.success("Servidor updated com sucesso! Empenhos vinculados foram sincronizados.")
                            except sqlite3.IntegrityError:
                                st.error("Ocorreu um conflito (CPF duplicado).")
                            finally:
                                conn.close()
                                st.rerun()
                                
                with c_deletar:
                    if st.button("❌ Excluir Servidor", type="secondary", key="btn_excluir_servidor"):
                        conn, cursor = conectar_banco()
                        try:
                            cursor.execute("SELECT COUNT(*) FROM diarias WHERE servidor_id = ?", (id_atual,))
                            possui_diarias = cursor.fetchone()[0]
                            
                            if possui_diarias > 0:
                                st.error("Não é possível excluir este servidor pois ele já possui diárias lançadas no módulo DEAEV.")
                            else:
                                cursor.execute("DELETE FROM servidores WHERE id = ?", (id_atual,))
                                conn.commit()
                                st.success("Servidor removido com sucesso!")
                        except Exception as e:
                            st.error(f"Erro ao remover: {e}")
                        finally:
                            conn.close()
                            st.rerun()

    # =====================================================================
    # MÓDULO EMPENHOS (CORREÇÃO DE SQL APLICADA)
    # =====================================================================
    elif st.session_state.menu_ativo == "empenhos":
        st.markdown("<h2>📝 Gerenciamento de Notas de Empenho</h2>", unsafe_allow_html=True)
        st.caption("⚠️ Modo Desenvolvimento: A trava de segurança foi suspensa. Excluir um empenho removerá automaticamente suas diárias.")
        
        if not lista_servidores:
            st.warning("⚠️ Atenção: Não há servidores cadastrados no sistema.")
        else:
            conn, cursor = conectar_banco()
            cursor.execute("SELECT id, numero_empenho, servidor_nome, valor_total, valor_disponivel, data_empenho, caminho_pdf FROM empenhos ORDER BY id DESC")
            empenhos_gravados = cursor.fetchall()
            conn.close()
            
            col_cad_e, col_cons_e = st.columns([1.1, 1.2])
            empenho_selecionado_df = None
            
            with col_cons_e:
                st.markdown("#### 🔍 Consulta de Empenhos")
                st.caption("Clique no círculo ao lado esquerdo do número do Empenho para editar ou excluí-lo.")
                
                if empenhos_gravados:
                    df_empenhos_raw = pd.DataFrame(empenhos_gravados, columns=["ID", "Número Empenho", "Servidor Vinculado", "Valor Inicial Raw", "Saldo Disponível Raw", "Data Raw", "Caminho PDF"])
                    
                    df_exibicao_e = df_empenhos_raw[["Número Empenho", "Servidor Vinculado"]].copy()
                    df_exibicao_e["Saldo Remanescente"] = df_empenhos_raw["Saldo Disponível Raw"].apply(lambda v: f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
                    
                    tabela_interativa_e = st.dataframe(
                        df_exibicao_e, 
                        use_container_width=True, 
                        hide_index=True, 
                        height=400,
                        on_select="rerun",
                        selection_mode="single-row",
                        key="tabela_empenhos_interativa"
                    )
                    
                    indices_selecionados_e = tabela_interativa_e.get("selection", {}).get("rows", [])
                    if indices_selecionados_e:
                        index_linha_e = indices_selecionados_e[0]
                        empenho_selecionado_df = df_empenhos_raw.iloc[index_linha_e]
                else:
                    st.info("Nenhum empenho cadastrado.")
                    
            with col_cad_e:
                if empenho_selecionado_df is not None:
                    st.markdown("#### ✏️ Alterar / Excluir Empenho")
                    id_emp_atual = int(empenho_selecionado_df["ID"])
                    num_emp_padrao = str(empenho_selecionado_df["Número Empenho"])
                    val_emp_padrao = float(empenho_selecionado_df["Valor Inicial Raw"])
                    serv_vinc_padrao = str(empenho_selecionado_df["Servidor Vinculado"])
                    
                    dt_str_ori = str(empenho_selecionado_df["Data Raw"])
                    try:
                        dt_padrao = datetime.strptime(dt_str_ori, "%Y-%m-%d").date()
                    except:
                        dt_padrao = datetime.now().date()
                else:
                    st.markdown("#### 🤖 Importação Automática por PDF")
                    arquivo_subido = st.file_uploader("Arraste ou selecione a Nota de Empenho em PDF do Paraná", type=["pdf"])
                    
                    if arquivo_subido is not None:
                        dados_extraidos = extrair_dados_pdf_pr(arquivo_subido)
                        if dados_extraidos:
                            st.session_state.pdf_num_empenho = dados_extraidos["empenho"]
                            st.session_state.pdf_data_empenho = dados_extraidos["data"]
                            st.session_state.pdf_valor_empenho = dados_extraidos["valor"]
                            
                            servidor_encontrado_label = ""
                            for s in lista_servidores:
                                if s["label"] in dados_extraidos["servidor"]:
                                    servidor_encontrado_label = s["label"]
                                    break
                            st.session_state.pdf_nome_servidor = servidor_encontrado_label
                            st.success("⚡ Dados do PDF carregados no formulário!")
                    
                    st.markdown("---")
                    st.markdown("#### ➕ Cadastrar Novo Empenho")
                    id_emp_atual = None
                    num_emp_padrao = st.session_state.pdf_num_empenho
                    val_emp_padrao = float(st.session_state.pdf_valor_empenho) if st.session_state.pdf_valor_empenho else 0.0
                    serv_vinc_padrao = st.session_state.pdf_nome_servidor
                    dt_padrao = st.session_state.pdf_data_empenho

                # Campos do Formulário de Empenho
                data_selecionada_e = st.date_input("Data de Emissão", value=dt_padrao, format="DD/MM/YYYY")
                
                # Lista de opções vinda dos servidores cadastrados
                opcoes_serv_emp = [s["label"] for s in lista_servidores]
                lista_chaves_servidores = [""] + opcoes_serv_emp
                
                try:
                    idx_selecionado_e = lista_chaves_servidores.index(serv_vinc_padrao)
                except ValueError:
                    idx_selecionado_e = 0
                    
                servidor_responsavel_e = st.selectbox("Nome do Servidor Vinculado", options=lista_chaves_servidores, index=idx_selecionado_e)
                numero_empenho_input = st.text_input("Número do Empenho", value=num_emp_padrao)
                valor_empenho_input = st.number_input("Valor do Empenho (R$)", min_value=0.0, value=val_emp_padrao, format="%.2f")
                
                # Regras de botões para Cadastrar ou Alterar/Excluir
                if id_emp_atual is None:
                    if st.button("➕ Salvar Empenho", type="primary", use_container_width=True):
                        if not servidor_responsavel_e or not numero_empenho_input.strip() or valor_empenho_input <= 0:
                            st.error("❌ Por favor, preencha todos os campos obrigatórios (incluindo o servidor).")
                        else:
                            num_emp_limpo = numero_empenho_input.strip().upper()
                            caminho_salvar_pdf = ""
                            
                            if 'arquivo_subido' in locals() and arquivo_subido is not None:
                                nome_arquivo_fisico = f"{num_emp_limpo}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
                                caminho_salvar_pdf = os.path.join(PASTA_EMPENHOS, nome_arquivo_fisico)
                                with open(caminho_salvar_pdf, "wb") as f:
                                    f.write(arquivo_subido.getbuffer())
                            
                            conn, cursor = conectar_banco()
                            try:
                                cursor.execute("""
                                    INSERT INTO empenhos (servidor_nome, numero_empenho, valor_total, valor_disponivel, data_empenho, caminho_pdf)
                                    VALUES (?, ?, ?, ?, ?, ?)
                                """, (servidor_responsavel_e, num_emp_limpo, valor_empenho_input, valor_empenho_input, str(data_selecionada_e), caminho_salvar_pdf))
                                conn.commit()
                                st.success(f"✔️ Empenho '{num_emp_limpo}' cadastrado!")
                                st.session_state.pdf_num_empenho = ""
                                st.session_state.pdf_data_empenho = datetime.now().date()
                                st.session_state.pdf_nome_servidor = ""
                                st.session_state.pdf_valor_empenho = 0.0
                            except sqlite3.IntegrityError:
                                st.error("❌ Este número de empenho já existe.")
                            finally:
                                conn.close()
                                st.rerun()
                else:
                    col_act_e, col_del_e = st.columns(2)
                    with col_act_e:
                        if st.button("💾 Atualizar Empenho", type="primary", use_container_width=True, key="btn_atualizar_emp_ok"):
                            if not servidor_responsavel_e or not numero_empenho_input.strip() or valor_empenho_input <= 0:
                                st.error("❌ Preencha todos os campos obrigatórios.")
                            else:
                                num_emp_limpo = numero_empenho_input.strip().upper()
                                conn, cursor = conectar_banco()
                                try:
                                    cursor.execute("SELECT valor_total, valor_disponivel FROM empenhos WHERE id = ?", (id_emp_atual,))
                                    v_tot_antigo, v_disp_antigo = cursor.fetchone()
                                    diferenca = valor_empenho_input - v_tot_antigo
                                    novo_disponivel = v_disp_antigo + diferenca
                                    
                                    if novo_disponivel < 0:
                                        st.error("❌ Não é possível reduzir o valor porque o saldo ficaria negativo pelas diárias já pagas.")
                                    else:
                                        cursor.execute("""
                                            UPDATE empenhos 
                                            SET servidor_nome = ?, numero_empenho = ?, valor_total = ?, valor_disponivel = ?, data_empenho = ?
                                            WHERE id = ?
                                        """, (servidor_responsavel_e, num_emp_limpo, valor_empenho_input, novo_disponivel, str(data_selecionada_e), id_emp_atual))
                                        sincronizar_pasta_empenho(cursor, id_emp_atual)
                                        conn.commit()
                                        st.success("✔️ Empenho atualizado com sucesso!")
                                        st.rerun()
                                except sqlite3.IntegrityError:
                                    st.error("❌ Conflito: Esse número de empenho já existe em outro registro.")
                                finally:
                                    conn.close()
                                    
                    with col_del_e:
                        if st.button("❌ Excluir Empenho", type="secondary", use_container_width=True, key="btn_excluir_emp_ok"):
                            conn, cursor = conectar_banco()
                            try:
                                # CORREÇÃO EXECUTADA AQUI: Alterado de "diárias" para "diarias"
                                cursor.execute("DELETE FROM diarias WHERE empenho_id = ?", (id_emp_atual,))
                                cursor.execute("DELETE FROM empenhos WHERE id = ?", (id_emp_atual,))
                                conn.commit()
                                st.success("✔️ Empenho e todas as suas diárias associadas foram removidos!")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Erro ao remover: {e}")
                            finally:
                                conn.close()

        st.markdown("---")
        st.markdown("#### 🗑️ Limpeza de Empenhos Zerados")
        st.caption("Quando o saldo de um empenho chega a zero, o PDF dele é movido automaticamente para a subpasta **EMPENHOS ZERADOS**, dentro da pasta de empenhos. Use o botão abaixo para excluir definitivamente esses arquivos e liberar espaço.")

        try:
            arquivos_zerados = [
                f for f in os.listdir(PASTA_EMPENHOS_ZERADOS)
                if os.path.isfile(os.path.join(PASTA_EMPENHOS_ZERADOS, f))
            ]
        except Exception:
            arquivos_zerados = []

        if not arquivos_zerados:
            st.info("Nenhum arquivo de empenho zerado aguardando exclusão no momento.")
        else:
            st.warning(f"📄 Existem **{len(arquivos_zerados)}** arquivo(s) de empenho(s) zerado(s) aguardando exclusão.")
            with st.expander("Ver arquivos que serão excluídos"):
                st.dataframe(pd.DataFrame(arquivos_zerados, columns=["Arquivo"]), use_container_width=True, hide_index=True)

            confirmar_exclusao_zerados = st.checkbox("Confirmo que desejo excluir permanentemente esses arquivos.", key="chk_confirma_exclusao_zerados")
            if st.button("🗑️ Excluir Empenhos Zerados", type="secondary", use_container_width=True, disabled=not confirmar_exclusao_zerados):
                conn, cursor = conectar_banco()
                total_excluidos = 0
                try:
                    for nome_arquivo in arquivos_zerados:
                        caminho_completo = os.path.join(PASTA_EMPENHOS_ZERADOS, nome_arquivo)
                        try:
                            os.remove(caminho_completo)
                            total_excluidos += 1
                            # Remove a referência do arquivo excluído no banco (o empenho em si permanece, preservando o histórico)
                            cursor.execute("UPDATE empenhos SET caminho_pdf = '' WHERE caminho_pdf = ?", (caminho_completo,))
                        except Exception:
                            pass
                    conn.commit()
                    st.success(f"✔️ {total_excluidos} arquivo(s) excluído(s) com sucesso da pasta de empenhos zerados!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Erro ao excluir arquivos: {e}")
                finally:
                    conn.close()

    # =====================================================================
    # MÓDULO DEAEV
    # =====================================================================
    elif st.session_state.menu_ativo == "deaev":
        st.markdown("<h2>📋 Registro de Diárias - DEAEV</h2>", unsafe_allow_html=True)
        
        if not lista_servidores:
            st.warning("⚠️ Atenção: Não há servidores cadastrados no sistema.")
        else:
            # Primeiro, carregamos as diárias do banco para mapear a seleção
            conn, cursor = conectar_banco()
            cursor.execute("""
                SELECT 
                    d.id, 
                    s.nome_completo AS servidor, 
                    s.id AS servidor_id,
                    e.numero_empenho, 
                    e.id AS empenho_id,
                    d.data_diaria, 
                    d.jornada, 
                    d.valor_diaria 
                FROM diarias d
                JOIN servidores s ON d.servidor_id = s.id
                JOIN empenhos e ON d.empenho_id = e.id
                ORDER BY d.data_diaria DESC, d.id DESC
            """)
            diarias_gravadas = cursor.fetchall()
            conn.close()

            diaria_selecionada_df = None

            # Dividindo a tela em duas colunas (Esquerda: Cadastro/Edição | Direita: Filtro e Lista)
            col_cadastro, col_lista = st.columns([1.2, 1.3])
            
            # ==========================================
            # COLUNA DA DIREITA: LISTA E FILTRO DE DIÁRIAS
            # ==========================================
            with col_lista:
                st.markdown("### 🔍 Consultar Lançamentos")
                st.caption("Clique no círculo ao lado esquerdo da diária para alterar ou excluir o registro.")
                
                meses_nomes = {
                    "01": "Janeiro", "02": "Fevereiro", "03": "Março", "04": "Abril", 
                    "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto", 
                    "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
                }
                
                col_filtro_mes, col_filtro_ano = st.columns(2)
                with col_filtro_mes:
                    mes_selecionado_nome = st.selectbox(
                        "Filtrar por Mês", 
                        options=["Todos"] + list(meses_nomes.values()),
                        index=0
                    )
                with col_filtro_ano:
                    anos_disponiveis = [str(ano) for ano in range(2023, 2030)]
                    ano_selecionado = st.selectbox(
                        "Filtrar por Ano", 
                        options=anos_disponiveis,
                        index=anos_disponiveis.index(str(datetime.now().year))
                    )
                
                if diarias_gravadas:
                    df_diarias = pd.DataFrame(diarias_gravadas, columns=[
                        "ID", "Servidor", "Servidor ID", "Empenho Num", "Empenho ID", "Data Raw", "Jornada", "Valor Raw"
                    ])
                    
                    df_diarias["Data Obj"] = pd.to_datetime(df_diarias["Data Raw"])
                    df_diarias["Mês Num"] = df_diarias["Data Obj"].dt.strftime("%m")
                    df_diarias["Ano Num"] = df_diarias["Data Obj"].dt.strftime("%Y")
                    df_diarias["Data Formated"] = df_diarias["Data Obj"].dt.strftime("%d/%m/%Y")
                    
                    df_filtrado = df_diarias[df_diarias["Ano Num"] == ano_selecionado]
                    if mes_selecionado_nome != "Todos":
                        mes_chave = [k for k, v in meses_nomes.items() if v == mes_selecionado_nome][0]
                        df_filtrado = df_filtrado[df_filtrado["Mês Num"] == mes_chave]
                    
                    if not df_filtrado.empty:
                        df_exibicao = pd.DataFrame()
                        df_exibicao["Servidor"] = df_filtrado["Servidor"]
                        df_exibicao["Data"] = df_filtrado["Data Formated"]
                        df_exibicao["Jornada"] = df_filtrado["Jornada"]
                        df_exibicao["Empenho"] = df_filtrado["Empenho Num"]
                        df_exibicao["Valor"] = df_filtrado["Valor Raw"].apply(lambda v: f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
                        
                        total_filtrado = df_filtrado["Valor Raw"].sum()
                        st.metric(
                            label=f"Total no período ({mes_selecionado_nome}/{ano_selecionado})", 
                            value=f"R$ {total_filtrado:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                        )
                        
                        tabela_diarias = st.dataframe(
                            df_exibicao, 
                            use_container_width=True, 
                            hide_index=True, 
                            height=350,
                            on_select="rerun",
                            selection_mode="single-row"
                        )
                        
                        indices_selecionados = tabela_diarias.get("selection", {}).get("rows", [])
                        if indices_selecionados:
                            index_linha = indices_selecionados[0]
                            diaria_selecionada_df = df_filtrado.iloc[index_linha]
                    else:
                        st.info(f"Nenhum lançamento encontrado para {mes_selecionado_nome}/{ano_selecionado}.")
                else:
                    st.info("Nenhuma diária lançada no sistema até o momento.")

            # ==========================================
            # COLUNA DA ESQUERDA: CADASTRO OU EDIÇÃO
            # ==========================================
            with col_cadastro:
                if diaria_selecionada_df is not None:
                    st.markdown("### ✏️ Alterar / Excluir Diária")
                    id_diaria_atual = int(diaria_selecionada_df["ID"])
                    id_empenho_original = int(diaria_selecionada_df["Empenho ID"])
                    valor_original = float(diaria_selecionada_df["Valor Raw"])
                    servidor_nome_original = str(diaria_selecionada_df["Servidor"])
                    data_original = datetime.strptime(str(diaria_selecionada_df["Data Raw"]), "%Y-%m-%d").date()
                    jornada_original = str(diaria_selecionada_df["Jornada"])
                else:
                    st.markdown("### 🆕 Nova Diária")
                    id_diaria_atual = None
                    servidor_nome_original = ""
                    data_original = datetime.today()
                    jornada_original = "6 Horas"
                
                # 1. Seleção do Servidor
                options_serv = {s["label"]: s["id"] for s in lista_servidores}
                lista_nomes_servidores = list(options_serv.keys())
                
                idx_serv_padrao = 0
                if servidor_nome_original in lista_nomes_servidores:
                    idx_serv_padrao = lista_nomes_servidores.index(servidor_nome_original) + 1
                
                servidor_selecionado = st.selectbox(
                    "Nome do Servidor", 
                    options=[""] + lista_nomes_servidores,
                    index=idx_serv_padrao,
                    disabled=(id_diaria_atual is not None)
                )
                
                # 2. Busca dinâmica de empenhos
                box_empenhos_filtrados = []
                if servidor_selecionado:
                    conn, cursor = conectar_banco()
                    if id_diaria_atual is not None:
                        cursor.execute("""
                            SELECT id, numero_empenho, valor_disponivel 
                            FROM empenhos 
                            WHERE servidor_nome = ? AND (valor_disponivel > 0 OR id = ?)
                            ORDER BY id DESC
                        """, (servidor_selecionado, id_empenho_original))
                    else:
                        cursor.execute("""
                            SELECT id, numero_empenho, valor_disponivel 
                            FROM empenhos 
                            WHERE servidor_nome = ? AND valor_disponivel > 0 
                            ORDER BY id DESC
                        """, (servidor_selecionado,))
                    
                    rows = cursor.fetchall()
                    conn.close()
                    
                    box_empenhos_filtrados = [
                        {"id": row[0], "label": f"Empenho {row[1]} (Saldo: R$ {row[2]:,.2f})".replace(",", "X").replace(".", ",").replace("X", ".")}
                        for row in rows
                    ]
                
                # 3. Exibição do selectbox de empenhos
                empenho_selecionado = None
                options_emp = {}
                if servidor_selecionado:
                    if not box_empenhos_filtrados:
                        st.error(f"❌ Erro: Não existem empenhos com saldo para {servidor_selecionado}.")
                    else:
                        options_emp = {e["label"]: e["id"] for e in box_empenhos_filtrados}
                        lista_labels_emp = list(options_emp.keys())
                        
                        idx_emp_padrao = 0
                        if id_diaria_atual is not None:
                            for idx, emp_item in enumerate(box_empenhos_filtrados):
                                if emp_item["id"] == id_empenho_original:
                                    idx_emp_padrao = idx + 1
                                    break
                        
                        empenho_selecionado = st.selectbox(
                            "Empenho Destinado", 
                            options=[""] + lista_labels_emp,
                            index=idx_emp_padrao
                        )
                else:
                    st.info("💡 Selecione um servidor para visualizar os empenhos.")
                    st.selectbox("Empenho Destinado", options=[""], disabled=True)
                
                # 4. Campos de Jornada e Valor Bloqueado
                c_data, c_jornada = st.columns(2)
                with c_data:
                    data_diaria = st.date_input("Data do Plantão", value=data_original, format="DD/MM/YYYY")
                with c_jornada:
                    idx_jornada = 0 if jornada_original == "6 Horas" else 1
                    jornada_selecionada = st.selectbox("Extra Jornada", options=["6 Horas", "12 Horas"], index=idx_jornada)
                
                # Cálculo automático do valor (Mantido travado para manter a consistência financeira)
                valor_calculado = 180.00 if jornada_selecionada == "6 Horas" else 360.00
                st.text_input("Valor da Diária", value=f"R$ {valor_calculado:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."), disabled=True)
                
                # 5. Processamento (Botões de Ação)
                if id_diaria_atual is None:
                    # MODO DE CADASTRO
                    if st.button("💾 Salvar Registro de Diária", type="primary", use_container_width=True):
                        if not servidor_selecionado or not empenho_selecionado:
                            st.error("❌ Por favor, selecione tanto o servidor quanto um empenho válido.")
                        else:
                            id_serv = options_serv[servidor_selecionado]
                            id_emp = options_emp[empenho_selecionado]
                            data_formatada_db = data_diaria.strftime("%Y-%m-%d")
                            
                            # Converter jornada atual para horas (inteiro) para validação
                            horas_novas = 6 if jornada_selecionada == "6 Horas" else 12
                            
                            conn, cursor = conectar_banco()
                            try:
                                # VALIDAÇÃO 1: Limite máximo de 12 horas por data
                                cursor.execute("""
                                    SELECT SUM(CASE WHEN jornada = '6 Horas' THEN 6 ELSE 12 END) 
                                    FROM diarias 
                                    WHERE servidor_id = ? AND data_diaria = ?
                                """, (id_serv, data_formatada_db))
                                horas_existentes = cursor.fetchone()[0] or 0
                                
                                if horas_existentes + horas_novas > 12:
                                    st.error(f"❌ Limite de Horas Excedido! O servidor já possui {horas_existentes} horas registradas em {data_diaria.strftime('%d/%m/%Y')}. Limite máximo é de 12 horas por dia.")
                                else:
                                    # VALIDAÇÃO 2: Saldo do empenho
                                    cursor.execute("SELECT valor_disponivel FROM empenhos WHERE id = ?", (id_emp,))
                                    saldo_atual = cursor.fetchone()[0]
                                    
                                    if saldo_atual >= valor_calculado:
                                        cursor.execute("""
                                            INSERT INTO diarias (servidor_id, empenho_id, data_diaria, jornada, valor_diaria) 
                                            VALUES (?, ?, ?, ?, ?)
                                        """, (id_serv, id_emp, data_formatada_db, jornada_selecionada, valor_calculado))
                                        
                                        cursor.execute("UPDATE empenhos SET valor_disponivel = valor_disponivel - ? WHERE id = ?", (valor_calculado, id_emp))
                                        sincronizar_pasta_empenho(cursor, id_emp)
                                        conn.commit()
                                        st.success("✔️ Diária registrada com sucesso!")
                                        st.rerun()
                                    else:
                                        st.error(f"❌ Operação Negada! Saldo insuficiente no empenho (Disponível: R$ {saldo_atual:,.2f}).")
                            except Exception as e:
                                conn.rollback()
                                st.error(f"Erro ao salvar: {e}")
                            finally:
                                conn.close()
                else:
                    # MODO DE ALTERAÇÃO / EXCLUSÃO
                    c_atualizar, c_excluir = st.columns(2)
                    
                    with c_atualizar:
                        if st.button("💾 Atualizar Diária", type="primary", use_container_width=True):
                            if not empenho_selecionado:
                                st.error("❌ Selecione um empenho válido.")
                            else:
                                id_emp_novo = options_emp[empenho_selecionado]
                                id_serv = int(diaria_selecionada_df["Servidor ID"])
                                data_formatada_db = data_diaria.strftime("%Y-%m-%d")
                                
                                # Converter jornada atual para horas para validação
                                horas_novas = 6 if jornada_selecionada == "6 Horas" else 12
                                
                                conn, cursor = conectar_banco()
                                try:
                                    # VALIDAÇÃO 1: Limite máximo de 12 horas por data (desconsiderando a própria diária atual que está sendo editada)
                                    cursor.execute("""
                                        SELECT SUM(CASE WHEN jornada = '6 Horas' THEN 6 ELSE 12 END) 
                                        FROM diarias 
                                        WHERE servidor_id = ? AND data_diaria = ? AND id != ?
                                    """, (id_serv, data_formatada_db, id_diaria_atual))
                                    horas_existentes = cursor.fetchone()[0] or 0
                                    
                                    if horas_existentes + horas_novas > 12:
                                        st.error(f"❌ Limite de Horas Excedido! Com esta alteração, o servidor somaria {horas_existentes + horas_novas} horas no dia {data_diaria.strftime('%d/%m/%Y')}. Limite máximo é de 12 horas por dia.")
                                    else:
                                        # Devolvemos primeiro o valor gasto anteriormente para recalcular o saldo temporariamente
                                        cursor.execute("UPDATE empenhos SET valor_disponivel = valor_disponivel + ? WHERE id = ?", (valor_original, id_empenho_original))
                                        
                                        # Checamos o saldo atualizado do empenho de destino
                                        cursor.execute("SELECT valor_disponivel FROM empenhos WHERE id = ?", (id_emp_novo,))
                                        saldo_disponivel_ajustado = cursor.fetchone()[0]
                                        
                                        # VALIDAÇÃO 2: Saldo de Empenho
                                        if saldo_disponivel_ajustado >= valor_calculado:
                                            cursor.execute("""
                                                UPDATE diarias 
                                                SET empenho_id = ?, data_diaria = ?, jornada = ?, valor_diaria = ?
                                                WHERE id = ?
                                            """, (id_emp_novo, data_formatada_db, jornada_selecionada, valor_calculado, id_diaria_atual))
                                            
                                            cursor.execute("UPDATE empenhos SET valor_disponivel = valor_disponivel - ? WHERE id = ?", (valor_calculado, id_emp_novo))
                                            sincronizar_pasta_empenho(cursor, id_empenho_original)
                                            sincronizar_pasta_empenho(cursor, id_emp_novo)
                                            conn.commit()
                                            st.success("✔️ Registro de diária updated com sucesso!")
                                            st.rerun()
                                        else:
                                            # Se faltou saldo, desfazemos o estorno que simulamos no início
                                            cursor.execute("UPDATE empenhos SET valor_disponivel = valor_disponivel - ? WHERE id = ?", (valor_original, id_empenho_original))
                                            sincronizar_pasta_empenho(cursor, id_empenho_original)
                                            conn.commit()
                                            st.error(f"❌ Operação Negada! O empenho selecionado não possui saldo suficiente (Disponível com estorno: R$ {saldo_disponivel_ajustado:,.2f}).")
                                except Exception as e:
                                    conn.rollback()
                                    st.error(f"Erro ao atualizar: {e}")
                                finally:
                                    conn.close()
                                    
                    with c_excluir:
                        if st.button("❌ Excluir Diária", type="secondary", use_container_width=True):
                            conn, cursor = conectar_banco()
                            try:
                                cursor.execute("UPDATE empenhos SET valor_disponivel = valor_disponivel + ? WHERE id = ?", (valor_original, id_empenho_original))
                                sincronizar_pasta_empenho(cursor, id_empenho_original)
                                cursor.execute("DELETE FROM diarias WHERE id = ?", (id_diaria_atual,))
                                conn.commit()
                                st.success("✔️ Registro removido e saldo estornado!")
                                st.rerun()
                            except Exception as e:
                                conn.rollback()
                                st.error(f"Erro ao excluir: {e}")
                            finally:
                                conn.close()

    # =====================================================================
    # MÓDULO RELATÓRIOS 📊
    # =====================================================================
    elif st.session_state.menu_ativo == "relatorios":
        st.markdown("<h2>📊 Relatórios e Prestação de Contas DEAEV</h2>", unsafe_allow_html=True)
        
        col_filtro_m, col_filtro_a, col_atualizar = st.columns([2, 2, 1])
        with col_filtro_m:
            mes_relatorio = st.selectbox("Mês de Competência", options=[f"{i:02d}" for i in range(1, 13)], index=datetime.now().month - 1)
        with col_filtro_a:
            ano_relatorio = st.selectbox("Ano de Competência", options=[str(a) for a in range(2024, 2031)], index=2)
        with col_atualizar:
            st.markdown("<div style='height: 28px;'></div>", unsafe_allow_html=True)
            btn_recarregar = st.button("🔄 Atualizar", use_container_width=True)
            
        competencia_busca = f"{ano_relatorio}-{mes_relatorio}"
        
        # Limpa o cache se o botão de atualizar for clicado
        if btn_recarregar and "dados_diarias_mes" in st.session_state:
            del st.session_state["dados_diarias_mes"]
        
        # Otimização: Busca no banco se os dados não existirem ou se a competência mudou
        if "dados_diarias_mes" not in st.session_state or st.session_state.get("competencia_atual") != competencia_busca:
            conn = None
            try:
                conn, cursor = conectar_banco()
                cursor.execute("""
                    SELECT s.unidade_lotacao, s.nome_completo, s.cpf, e.numero_empenho, d.data_diaria, d.jornada, d.valor_diaria
                    FROM diarias d
                    JOIN servidores s ON d.servidor_id = s.id
                    JOIN empenhos e ON d.empenho_id = e.id
                    WHERE strftime('%Y-%m', d.data_diaria) = ?
                    ORDER BY s.nome_completo ASC, e.numero_empenho ASC, d.data_diaria ASC
                """, (competencia_busca,))
                st.session_state.dados_diarias_mes = cursor.fetchall()
                st.session_state.competencia_atual = competencia_busca
            except Exception as e:
                st.error(f"Erro ao buscar dados do banco: {e}")
                st.session_state.dados_diarias_mes = []
            finally:
                if conn:
                    conn.close()
        
        dados_diarias_mes = st.session_state.dados_diarias_mes
        
        if not dados_diarias_mes:
            st.info(f"Nenhum registro de diária (DEAEV) encontrado para {mes_relatorio}/{ano_relatorio}.")
        else:
            st.success(f"Foram encontradas **{len(dados_diarias_mes)}** diárias lançadas neste mês.")
            st.markdown("---")
            
            if "relatorio_tipo_selecionado" not in st.session_state:
                st.session_state.relatorio_tipo_selecionado = None
                
            # Divisão equilibrada para os quatro botões principais
            col_btn1, col_btn2, col_btn3, col_btn4 = st.columns(4)
            with col_btn1:
                tipo_btn_deaev = "primary" if st.session_state.relatorio_tipo_selecionado == "deaev_realizada" else "secondary"
                if st.button("📋 DEAEV REALIZADA", type=tipo_btn_deaev, use_container_width=True):
                    st.session_state.relatorio_tipo_selecionado = "deaev_realizada"
                    st.rerun()
            with col_btn2:
                tipo_btn_template = "primary" if st.session_state.relatorio_tipo_selecionado == "template_final" else "secondary"
                if st.button("🏢 TEMPLATE FINAL", type=tipo_btn_template, use_container_width=True):
                    st.session_state.relatorio_tipo_selecionado = "template_final"
                    st.rerun()
            with col_btn3:
                tipo_btn_servidores = "primary" if st.session_state.relatorio_tipo_selecionado == "servidores" else "secondary"
                if st.button("👤 SERVIDORES", type=tipo_btn_servidores, use_container_width=True):
                    st.session_state.relatorio_tipo_selecionado = "servidores"
                    st.rerun()
            with col_btn4:
                tipo_btn_empenhos_utilizados = "primary" if st.session_state.relatorio_tipo_selecionado == "empenhos_utilizados" else "secondary"
                if st.button("📦 EMPENHOS UTILIZADOS", type=tipo_btn_empenhos_utilizados, use_container_width=True):
                    st.session_state.relatorio_tipo_selecionado = "empenhos_utilizados"
                    st.rerun()
            
            st.markdown("<br>", unsafe_allow_html=True)
            
            # --- FLUXO DO BOTÃO 1: DEAEV REALIZADA ---
            if st.session_state.relatorio_tipo_selecionado == "deaev_realizada":
                st.markdown("### 🔍 Opções para: **DEAEV REALIZADA**")
                formato_deaev = st.radio("Escolha o formato de saída:", options=["PDF", "PLANILHA"], index=0, horizontal=True)
                
                st.markdown("---")
                num_processo_op = st.text_input("Informe o Número do Processo / Protocolo:", placeholder="Ex: 26.216.461-6", key="proc_deaev")
                
                if num_processo_op.strip() == "":
                    st.warning("⚠️ Digite o número do processo acima para liberar a geração do relatório.")
                else:
                    titulo_tabela = f"TABELA 01 - PAGAMENTO {MESES_EXTENSO[mes_relatorio]} / {ano_relatorio}"
                    titulo_lote = f"SESP/DEPPEN – PAGAMENTO DE EXTRAJORNADA – REF. {mes_relatorio}/{ano_relatorio} – REGIONAL DE UMUARAMA – PROTOCOLO {num_processo_op.strip()}"
                    
                    if formato_deaev == "PLANILHA":
                        wb_op = Workbook()
                        ws_op = wb_op.active
                        ws_op.title = "Worksheet"
                        
                        FONTE_XLSX = "Calibri"
                        COR_CABECALHO = "FF729FCD"
                        COR_BRANCO = "FFFFFFFF"
                        borda_fina = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
                        alinhamento_centro = Alignment(horizontal="center", vertical="center")
                        
                        titulo_tabela_xlsx = f"TABELA 01 – PAGAMENTO {MESES_EXTENSO[mes_relatorio]} {ano_relatorio}"
                        
                        ws_op.append([titulo_tabela_xlsx])
                        ws_op.merge_cells("A1:H1")
                        ws_op["A1"].font = Font(name=FONTE_XLSX, size=11, bold=True)
                        ws_op["A1"].alignment = alinhamento_centro
                        ws_op.row_dimensions[1].height = 30
                        for col in range(1, 9): 
                            ws_op.cell(row=1, column=col).border = borda_fina
                        
                        ws_op.append([titulo_lote])
                        ws_op.merge_cells("A2:H2")
                        ws_op["A2"].font = Font(name=FONTE_XLSX, size=11, bold=True)
                        ws_op["A2"].alignment = alinhamento_centro
                        ws_op.row_dimensions[2].height = 30
                        for col in range(1, 9): 
                            ws_op.cell(row=2, column=col).border = borda_fina
                        
                        headers_op = ["Unidade", "Policial", "CPF", "Empenho", "Data", "Horas", "Valor", "Total"]
                        ws_op.append(headers_op)
                        ws_op.row_dimensions[3].height = 30
                        for col_num in range(1, len(headers_op) + 1):
                            cell = ws_op.cell(row=3, column=col_num)
                            cell.font = Font(name=FONTE_XLSX, size=11, bold=True)
                            cell.fill = PatternFill(start_color=COR_CABECALHO, end_color=COR_CABECALHO, fill_type="solid")
                            cell.alignment = alinhamento_centro
                            cell.border = borda_fina
                        
                        larguras_colunas = {"A": 12.0, "B": 50.56, "C": 16.28, "D": 17.57, "E": 15.14, "F": 9.28, "G": 14.0, "H": 16.28}
                        for col_letra, largura in larguras_colunas.items():
                            ws_op.column_dimensions[col_letra].width = largura
                        
                        # Agrupamento em dois níveis:
                        # 1) Por SERVIDOR (Unidade/Policial/CPF mesclam mesmo que o servidor tenha diárias em empenhos diferentes,
                        #    evitando repetir o nome do mesmo servidor várias vezes no relatório)
                        # 2) Por EMPENHO dentro de cada servidor (Empenho/Total continuam se mesclando como já era)
                        grupos_servidor_xlsx = []
                        grupo_servidor_atual_xlsx = None
                        for row_data in dados_diarias_mes:
                            unidade, policial, cpf, empenho, data_db, jornada, valor = row_data
                            sigla_unidade = unidade.split(" - ")[0].strip() if unidade else ""
                            dt_br = datetime.strptime(data_db, "%Y-%m-%d").strftime("%d/%m/%Y")
                            num_horas = 6 if "6" in jornada else 12
                            
                            if grupo_servidor_atual_xlsx is not None and grupo_servidor_atual_xlsx["policial"] == policial:
                                subgrupo_atual_xlsx = grupo_servidor_atual_xlsx["subgrupos"][-1]
                                if subgrupo_atual_xlsx["empenho"] == empenho:
                                    subgrupo_atual_xlsx["linhas"].append((dt_br, num_horas, float(valor)))
                                else:
                                    grupo_servidor_atual_xlsx["subgrupos"].append({
                                        "empenho": empenho,
                                        "linhas": [(dt_br, num_horas, float(valor))]
                                    })
                            else:
                                grupo_servidor_atual_xlsx = {
                                    "unidade": sigla_unidade,
                                    "policial": policial,
                                    "cpf": cpf,
                                    "subgrupos": [{"empenho": empenho, "linhas": [(dt_br, num_horas, float(valor))]}]
                                }
                                grupos_servidor_xlsx.append(grupo_servidor_atual_xlsx)
                        
                        linha_atual_xlsx = 4
                        for grupo_serv in grupos_servidor_xlsx:
                            linha_inicio_servidor = linha_atual_xlsx
                            primeira_linha_servidor = True
                            
                            for subgrupo in grupo_serv["subgrupos"]:
                                qtd_linhas_grupo = len(subgrupo["linhas"])
                                total_grupo = sum(item[2] for item in subgrupo["linhas"])
                                linha_inicio_grupo = linha_atual_xlsx
                                linha_fim_grupo = linha_atual_xlsx + qtd_linhas_grupo - 1
                                
                                for idx, (dt_br, num_horas, valor) in enumerate(subgrupo["linhas"]):
                                    linha = linha_atual_xlsx + idx
                                    ws_op.cell(row=linha, column=1, value=grupo_serv["unidade"] if primeira_linha_servidor and idx == 0 else None)
                                    ws_op.cell(row=linha, column=2, value=grupo_serv["policial"] if primeira_linha_servidor and idx == 0 else None)
                                    ws_op.cell(row=linha, column=3, value=str(grupo_serv["cpf"]) if primeira_linha_servidor and idx == 0 else None)
                                    ws_op.cell(row=linha, column=4, value=subgrupo["empenho"] if idx == 0 else None)
                                    ws_op.cell(row=linha, column=5, value=dt_br)
                                    ws_op.cell(row=linha, column=6, value=num_horas)
                                    
                                    cell_val = ws_op.cell(row=linha, column=7, value=valor)
                                    cell_val.number_format = 'R$#,##0.00'
                                    
                                    if idx == 0:
                                        cell_tot = ws_op.cell(row=linha, column=8, value=total_grupo)
                                        cell_tot.number_format = 'R$#,##0.00'
                                    else:
                                        ws_op.cell(row=linha, column=8, value=None)
                                    
                                    for col_num in range(1, len(headers_op) + 1):
                                        cell = ws_op.cell(row=linha, column=col_num)
                                        cell.font = Font(name=FONTE_XLSX, size=11, bold=False)
                                        cell.fill = PatternFill(start_color=COR_BRANCO, end_color=COR_BRANCO, fill_type="solid")
                                        cell.alignment = alinhamento_centro
                                        cell.border = borda_fina
                                
                                if qtd_linhas_grupo > 1:
                                    for col_num_mesclar in [4, 8]:
                                        letra_col = get_column_letter(col_num_mesclar)
                                        ws_op.merge_cells(f"{letra_col}{linha_inicio_grupo}:{letra_col}{linha_fim_grupo}")
                                
                                linha_atual_xlsx = linha_fim_grupo + 1
                                primeira_linha_servidor = False
                            
                            linha_fim_servidor = linha_atual_xlsx - 1
                            if linha_fim_servidor > linha_inicio_servidor:
                                for col_num_mesclar in [1, 2, 3]:
                                    letra_col = get_column_letter(col_num_mesclar)
                                    ws_op.merge_cells(f"{letra_col}{linha_inicio_servidor}:{letra_col}{linha_fim_servidor}")
                        
                        total_policiais = len(set(item[1] for item in dados_diarias_mes))
                        total_horas_mes = sum(6 if "6" in item[5] else 12 for item in dados_diarias_mes)
                        total_deaev = total_horas_mes / 12
                        total_deaev_str = formatar_qtd_deaev(total_deaev)
                        total_geral = sum(float(item[6]) for item in dados_diarias_mes)
                        
                        textos_totais = [
                            f"Total de Policiais: {total_policiais}",
                            f"Total de DEAEV: {total_deaev_str}",
                            f"TOTAL: R$ {formatar_valor_br(total_geral)}"
                        ]
                        for texto_total in textos_totais:
                            ws_op.cell(row=linha_atual_xlsx, column=1, value=texto_total)
                            ws_op.merge_cells(f"A{linha_atual_xlsx}:H{linha_atual_xlsx}")
                            
                            for col in range(1, 9):
                                cell_borda = ws_op.cell(row=linha_atual_xlsx, column=col)
                                cell_borda.border = borda_fina
                                cell_borda.fill = PatternFill(start_color=COR_CABECALHO, end_color=COR_CABECALHO, fill_type="solid")
                            
                            cell = ws_op.cell(row=linha_atual_xlsx, column=1)
                            cell.font = Font(name=FONTE_XLSX, size=11, bold=True)
                            cell.alignment = alinhamento_centro
                            ws_op.row_dimensions[linha_atual_xlsx].height = 20
                            linha_atual_xlsx += 1
                                    
                        caminho_temp_xlsx_op = os.path.join(DIRETORIO_ATUAL, "temp_op.xlsx")
                        wb_op.save(caminho_temp_xlsx_op)
                        with open(caminho_temp_xlsx_op, "rb") as f_op:
                            st.download_button(
                                "🟢 Baixar DEAEV REALIZADA em Excel (.xlsx)", 
                                data=f_op.read(), 
                                file_name=f"relatorio_mensal_{mes_relatorio}_{ano_relatorio}.xlsx", 
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                                use_container_width=True
                            )
                            
                    elif formato_deaev == "PDF":
                        caminho_temp_pdf_op = os.path.join(DIRETORIO_ATUAL, "temp_op.pdf")
                        doc_op = SimpleDocTemplate(caminho_temp_pdf_op, pagesize=landscape(letter), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
                        story_op = []
                        
                        estilos = getSampleStyleSheet()
                        estilo_titulo = ParagraphStyle('TitOp', parent=estilos['Heading3'], alignment=1, spaceAfter=4)
                        estilo_subtitulo = ParagraphStyle('SubTitOp', parent=estilos['Normal'], alignment=1, fontSize=9, spaceAfter=15)
                        story_op.append(Paragraph(titulo_tabela, estilo_titulo))
                        story_op.append(Paragraph(titulo_lote, estilo_subtitulo))
                        
                        headers_op = ["Policial Penal", "CPF", "Empenhos", "Data", "Horas", "Valor", "Total"]
                        tabela_dados_pdf = [headers_op]
                        
                        estilos_tabela = [
                            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0,0), (-1,-1), 9),
                            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                        ]
                        
                        # Agrupamento em dois níveis, igual ao Excel:
                        # 1) Por SERVIDOR (Policial/CPF mesclam mesmo com empenhos diferentes)
                        # 2) Por EMPENHO dentro do servidor (Empenho/Total continuam se mesclando como já era)
                        grupos_servidor_pdf = []
                        grupo_servidor_atual_pdf = None
                        
                        for row_data in dados_diarias_mes:
                            unidade, policial, cpf, empenho, data_db, jornada, valor = row_data
                            dt_br = datetime.strptime(data_db, "%Y-%m-%d").strftime("%d/%m/%Y")
                            num_horas = 6 if "6" in jornada else 12
                            
                            if grupo_servidor_atual_pdf is not None and grupo_servidor_atual_pdf["policial"] == policial:
                                subgrupo_atual_pdf = grupo_servidor_atual_pdf["subgrupos"][-1]
                                if subgrupo_atual_pdf["empenho"] == empenho:
                                    subgrupo_atual_pdf["linhas"].append((dt_br, num_horas, valor))
                                else:
                                    grupo_servidor_atual_pdf["subgrupos"].append({"empenho": empenho, "linhas": [(dt_br, num_horas, valor)]})
                            else:
                                grupo_servidor_atual_pdf = {
                                    "policial": policial,
                                    "cpf": cpf,
                                    "subgrupos": [{"empenho": empenho, "linhas": [(dt_br, num_horas, valor)]}]
                                }
                                grupos_servidor_pdf.append(grupo_servidor_atual_pdf)
                        
                        linha_atual_pdf = 1
                        for grupo_serv in grupos_servidor_pdf:
                            linha_inicio_servidor = linha_atual_pdf
                            
                            for subgrupo in grupo_serv["subgrupos"]:
                                qtd_linhas = len(subgrupo["linhas"])
                                total_grupo = sum(item[2] for item in subgrupo["linhas"])
                                
                                if qtd_linhas > 1:
                                    fim_mesclagem = 	linha_atual_pdf + qtd_linhas - 1
                                    estilos_tabela.append(('SPAN', (2, linha_atual_pdf), (2, fim_mesclagem)))
                                    estilos_tabela.append(('SPAN', (6, linha_atual_pdf), (6, fim_mesclagem)))
                                
                                for idx, (dt_br, num_horas, valor) in enumerate(subgrupo["linhas"]):
                                    tabela_dados_pdf.append([
                                        grupo_serv["policial"][:30],
                                        grupo_serv["cpf"],
                                        subgrupo["empenho"],
                                        dt_br,
                                        str(num_horas),
                                        formatar_valor_br(valor),
                                        formatar_valor_br(total_grupo)
                                    ])
                                
                                linha_atual_pdf += qtd_linhas
                            
                            linha_fim_servidor = linha_atual_pdf - 1
                            if linha_fim_servidor > linha_inicio_servidor:
                                estilos_tabela.append(('SPAN', (0, linha_inicio_servidor), (0, linha_fim_servidor)))
                                estilos_tabela.append(('SPAN', (1, linha_inicio_servidor), (1, linha_fim_servidor)))
                                    
                        t_op = Table(tabela_dados_pdf, colWidths=[220, 95, 95, 75, 50, 70, 70], repeatRows=1)
                        t_op.setStyle(TableStyle(estilos_tabela))
                        story_op.append(t_op)
                        
                        total_policiais = len(set(item[1] for item in dados_diarias_mes))
                        total_horas_mes = sum(6 if "6" in item[5] else 12 for item in dados_diarias_mes)
                        total_deaev = total_horas_mes / 12
                        total_deaev_str = formatar_qtd_deaev(total_deaev)
                        total_geral = sum(item[6] for item in dados_diarias_mes)
                        
                        story_op.append(Spacer(1, 15))
                        tabela_totais_pdf = [
                            ["Total de Policiais", "Total de DEAEV", "Total em R$"],
                            [str(total_policiais), total_deaev_str, formatar_valor_br(total_geral)]
                        ]
                        t_op_totais = Table(tabela_totais_pdf, colWidths=[220, 220, 220])
                        t_op_totais.setStyle(TableStyle([
                            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                            ('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold'),
                            ('FONTSIZE', (0,0), (-1,-1), 10),
                            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                        ]))
                        story_op.append(t_op_totais)
                        
                        doc_op.build(story_op)
                        
                        with open(caminho_temp_pdf_op, "rb") as f_pdf_op:
                            st.download_button(
                                "🔴 Baixar DEAEV REALIZADA em PDF (.pdf)", 
                                data=f_pdf_op.read(), 
                                file_name=f"relatorio_mensal_{mes_relatorio}_{ano_relatorio}.pdf", 
                                mime="application/pdf", 
                                use_container_width=True
                            )
            
            # --- FLUXO DO BOTÃO 2: TEMPLATE FINAL ---
            elif st.session_state.relatorio_tipo_selecionado == "template_final":
                st.markdown("### 🔍 Opções para: **TEMPLATE FINAL**")
                formato_template = st.radio("Escolha o formato de saída:", options=["PLANILHA", "PDF"], index=0, horizontal=True)
                
                st.markdown("---")
                num_processo_tp = st.text_input("Informe o Número do Processo / Protocolo:", placeholder="Ex: 26.216.461-6", key="proc_template")
                
                if num_processo_tp.strip() == "":
                    st.warning("⚠️ Digite o número do processo acima para liberar a geração do template.")
                else:
                    df_bruto = pd.DataFrame(dados_diarias_mes, columns=["Unidade", "Policial", "CPF", "Empenho", "Data", "Jornada", "Valor"])
                    df_consolidado = df_bruto.groupby(["Empenho", "CPF", "Policial"]).agg({"Valor": "sum"}).reset_index()
                    
                    import calendar
                    ultimo_dia = calendar.monthrange(int(ano_relatorio), int(mes_relatorio))[1]
                    data_emissao_nl = f"{ultimo_dia:02d}/{mes_relatorio}/{ano_relatorio}"
                    
                    obs_padrao = f"SESP/DEPPEN - PAGAMENTO DE EXTRAJORNADA – REF. {mes_relatorio}/{ano_relatorio} - REGIONAL DE UMUARAMA"
                    
                    headers_tp = [
                        'Data de Emissão NL', 'UG Emitente', 'Nota de Empenho', 'Tipo Patrimonial', 'Item Patrimonial', 
                        'Operação Patrimonial', 'Valor do Item', 'Tipo de Retenção', 'Credor da Retenção', 
                        'Valor da Base de Cálculo da Retenção', 'Valor da Retenção', 'Observação', 'Mês de Competência', 
                        'Data do Processo', 'Código do Processo', 'Credor Secundário', 'DEA', 'Tipo de Documento Comprobatório', 
                        'Número Documento Comprobatório', 'Processo Documento Comprobatório', 'Data Documento Comprobatório', 
                        'Competência Documento Comprobatório', 'Tipo de Série Documento Comprobatório', 
                        'Descrição Tipo de Série Documento Comprobatório', 'Chave de Acesso Documento Comprobatório', 
                        'Código de Verificação Documento Comprobatório', 'Valor Documento Comprobatório', 'Categoria do PADV'
                    ]
                    
                    larguras_colunas = {
                        'A': 9.29, 'B': 6.29, 'C': 11.29, 'D': 4.29, 'E': 4.29, 'F': 4.29,
                        'G': 8.29, 'H': 4.29, 'I': 4.29, 'J': 4.29, 'K': 4.29, 'L': 24.29,
                        'M': 4.29, 'N': 4.29, 'O': 13.29, 'P': 4.29, 'Q': 7.29, 'R': 4.29,
                        'S': 9.29, 'T': 13.29, 'U': 9.29, 'V': 9.29, 'W': 4.29, 'X': 4.29,
                        'Y': 4.29, 'Z': 4.29, 'AA': 12.29, 'AB': 11.29
                    }
                    
                    colunas_cinza = {2, 4, 5, 6, 17, 23} # Índices base 1 (B, D, E, F, Q, W)
                    colunas_sem_fundo_letras = {'H', 'I', 'J', 'K', 'N', 'P', 'X', 'Y', 'Z'}
                    
                    # ================= OPÇÃO 1: FORMATO PLANILHA (EXCEL) =================
                    if formato_template == "PLANILHA":
                        wb_tp = Workbook()
                        ws_tp = wb_tp.active
                        ws_tp.title = "Worksheet"
                        
                        COR_AZUL_CLARO = "4169E1"
                        COR_CINZA = "D3D3D3"
                        
                        fonte_cabecalho_branca = Font(name="Calibri", size=8, bold=False, color="FFFFFF")
                        fonte_cabecalho_preta = Font(name="Calibri", size=8, bold=False, color="000000")
                        fonte_dados = Font(name="Calibri", size=8, bold=False, color="000000")
                        
                        fill_azul = PatternFill(start_color=COR_AZUL_CLARO, end_color=COR_AZUL_CLARO, fill_type="solid")
                        fill_cinza = PatternFill(start_color=COR_CINZA, end_color=COR_CINZA, fill_type="solid")
                        
                        alinhamento_cabecalho = Alignment(horizontal="left", vertical="center", wrap_text=False)
                        alinhamento_dados = Alignment(horizontal="left", vertical="center", wrap_text=False)
                        
                        ws_tp.append(headers_tp)
                        ws_tp.row_dimensions[1].height = 20
                        
                        for col_letra, tamanho in larguras_colunas.items():
                            ws_tp.column_dimensions[col_letra].width = tamanho
                        
                        colunas_sem_fundo_idx = {8, 9, 10, 11, 14, 16, 24, 25, 26}
                        
                        for col_num in range(1, len(headers_tp) + 1):
                            cell = ws_tp.cell(row=1, column=col_num)
                            cell.alignment = alinhamento_cabecalho
                            
                            if col_num in colunas_sem_fundo_idx:
                                cell.fill = PatternFill(fill_type=None)
                                cell.font = fonte_cabecalho_preta
                            elif col_num in colunas_cinza:
                                cell.fill = fill_cinza
                                cell.font = fonte_cabecalho_preta
                            else:
                                cell.fill = fill_azul
                                cell.font = fonte_cabecalho_branca
                        
                        for idx, row_c in df_consolidado.iterrows():
                            ws_tp.append([
                                data_emissao_nl, 390000, row_c["Empenho"], 15, 1902, 58, row_c["Valor"], "", "", "", "",
                                obs_padrao, int(mes_relatorio), "", num_processo_tp.strip(), "", "00000000", 23, f"{mes_relatorio}/{ano_relatorio}",
                                num_processo_tp.strip(), data_emissao_nl, f"{mes_relatorio}/{ano_relatorio}", 6, "", "", "", row_c["Valor"], ""
                            ])
                            
                            linha_atual = ws_tp.max_row
                            ws_tp.row_dimensions[linha_atual].height = 16
                            
                            for col_num in range(1, len(headers_tp) + 1):
                                cell_dados = ws_tp.cell(row=linha_atual, column=col_num)
                                cell_dados.font = fonte_dados
                                cell_dados.alignment = alinhamento_dados
                            
                            ws_tp.cell(row=linha_atual, column=17).number_format = '@'
                            
                        caminho_temp_xlsx_tp = os.path.join(DIRETORIO_ATUAL, "temp_tp.xlsx")
                        wb_tp.save(caminho_temp_xlsx_tp)
                        with open(caminho_temp_xlsx_tp, "rb") as f_tp:
                            st.download_button("🟢 Baixar TEMPLATE FINAL em Excel (.xlsx)", data=f_tp.read(), file_name=f"TEMPLATE_{mes_relatorio}_{ano_relatorio}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                            
                    # ================= OPÇÃO 2: FORMATO PDF (FONTE EQUILIBRADA + DENSIDADE EXTREMA) =================
                    elif formato_template == "PDF":
                        import os
                        from reportlab.lib import colors
                        from reportlab.lib.pagesizes import letter, landscape
                        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
                        
                        caminho_temp_pdf_tp = os.path.join(DIRETORIO_ATUAL, "temp_tp.pdf")
                        
                        # Espaço horizontal total utilizável na página Letter Landscape com margens de 10pt = 772pt
                        LARGURA_UTIL_PAGINA = 772.0
                        
                        # Filtra e calcula a soma das proporções das colunas visíveis
                        total_unidades_visiveis = sum([larguras_colunas[letra] for i, letra in enumerate(larguras_colunas) if letra not in colunas_sem_fundo_letras and headers_tp[i] != 'Categoria do PADV'])
                        fator_escala = LARGURA_UTIL_PAGINA / total_unidades_visiveis
                        
                        # Dicionário estrito corrigido com 'Mês'
                        mapa_abreviações_exatas = {
                            'Data de Emissão NL': 'Data', 'UG Emitente': 'UG', 'Nota de Empenho': 'Not',
                            'Tipo Patrimonial': 'Tipo', 'Item Patrimonial': 'Item',
                            'Operação Patrimonial': 'Opi', 'Valor do Item': 'Val', 'Observação': 'Obs',
                            'Mês de Competência': 'Mês', 'Código do Processo': 'Cod', 'DEA': 'DEA',
                            'Tipo de Documento Comprobatório': 'Tip', 'Número Documento Comprobatório': 'Num',
                            'Processo Documento Comprobatório': 'Proc', 'Data Documento Comprobatório': 'Data',
                            'Competência Documento Comprobatório': 'Comp', 'Tipo de Série Documento Comprobatório': 'Cod',
                            'Valor Documento Comprobatório': 'Val'
                        }
                        
                        colunas_visiveis = []
                        for i, h_text in enumerate(headers_tp):
                            col_num = i + 1
                            letra = chr(ord('A') + i) if i < 26 else 'AA' if i == 26 else 'AB'
                            if letra not in colunas_sem_fundo_letras and h_text != 'Categoria do PADV':
                                titulo_reduzido = mapa_abreviações_exatas.get(h_text, h_text)
                                largura_calculada = larguras_colunas[letra] * fator_escala
                                
                                # Ajuste: Deixa a coluna do Mês 35% mais estreita
                                if titulo_reduzido == 'Mês':
                                    largura_calculada = largura_calculada * 0.65
                                
                                colunas_visiveis.append({
                                    'idx_0': i,
                                    'col_num': col_num,
                                    'letra': letra,
                                    'titulo': titulo_reduzido,
                                    'is_cinza': col_num in colunas_cinza,
                                    'width_pt': largura_calculada
                                })
                        
                        larguras_pdf_filtradas = [c['width_pt'] for c in colunas_visiveis]
                        
                        # Configuração do Documento
                        doc_tp = SimpleDocTemplate(
                            caminho_temp_pdf_tp, 
                            pagesize=landscape(letter), 
                            rightMargin=10, leftMargin=10, topMargin=15, bottomMargin=15
                        )
                        story_tp = []
                        
                        # --- CONSTRUÇÃO DOS DADOS ---
                        tabela_tp_pdf = []
                        
                        # 1. Cabeçalho
                        tabela_tp_pdf.append([c['titulo'] for c in colunas_visiveis])
                        
                        # 2. Linhas de Dados (Formatadas sem pontos decimais flutuantes)
                        for idx, row_c in df_consolidado.iterrows():
                            linha_completa = [
                                data_emissao_nl, 390000, row_c["Empenho"], 15, 1902, 58, int(row_c['Valor']), "", "", "", "",
                                obs_padrao, int(mes_relatorio), "", num_processo_tp.strip(), "", "00000000", 23, f"{mes_relatorio}/{ano_relatorio}",
                                num_processo_tp.strip(), data_emissao_nl, f"{mes_relatorio}/{ano_relatorio}", 6, "", "", "", int(row_c['Valor']), ""
                            ]
                            
                            linha_filtrada = []
                            for c in colunas_visiveis:
                                valor_celula = str(linha_completa[c['idx_0']])
                                if c['titulo'] == 'Obs' and len(valor_celula) > 30:
                                    valor_celula = valor_celula[:28] + ".."
                                linha_filtrada.append(valor_celula)
                                
                            tabela_tp_pdf.append(linha_filtrada)
                        
                        # --- ESTILIZAÇÃO COM AJUSTE CIRÚRGICO DE FONTE ---
                        estilos_tabela = [
                            ('ALIGN', (0,0), (-1,-1), 'CENTER'),    
                            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),   
                            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
                            ('FONTSIZE', (0,0), (-1,-1), 5.5),      # Aumentado sutilmente para melhor leitura
                            ('LEADING', (0,0), (-1,-1), 5.5),       # Acompanha o tamanho da fonte para manter o controle
                            
                            # Micro-ajuste compensatório no padding para a linha não estourar e manter o visual estreito
                            ('TOPPADDING', (0,0), (-1,-1), 0.8),    
                            ('BOTTOMPADDING', (0,0), (-1,-1), 0.8), 
                            
                            ('LEFTPADDING', (0,0), (-1,-1), 2),
                            ('RIGHTPADDING', (0,0), (-1,-1), 2),
                        ]
                        
                        # Identifica a coluna 'Mês' para aplicar o alinhamento exclusivo à esquerda
                        idx_mes = next((idx for idx, c in enumerate(colunas_visiveis) if c['titulo'] == 'Mês'), None)
                        if idx_mes is not None:
                            estilos_tabela.append(('ALIGN', (idx_mes, 0), (idx_mes, -1), 'LEFT'))
                            estilos_tabela.append(('LEFTPADDING', (idx_mes, 0), (idx_mes, -1), 4))
                        
                        # Cores dos Cabeçalhos
                        for idx_visivel, col_info in enumerate(colunas_visiveis):
                            if col_info['is_cinza']:
                                estilos_tabela.append(('BACKGROUND', (idx_visivel, 0), (idx_visivel, 0), colors.HexColor('#E2E8F0')))
                                estilos_tabela.append(('TEXTCOLOR', (idx_visivel, 0), (idx_visivel, 0), colors.HexColor('#1E293B')))
                                estilos_tabela.append(('FONTNAME', (idx_visivel, 0), (idx_visivel, 0), 'Helvetica-Bold'))
                            else:
                                estilos_tabela.append(('BACKGROUND', (idx_visivel, 0), (idx_visivel, 0), colors.HexColor('#1E3A8A')))
                                estilos_tabela.append(('TEXTCOLOR', (idx_visivel, 0), (idx_visivel, 0), colors.white))
                                estilos_tabela.append(('FONTNAME', (idx_visivel, 0), (idx_visivel, 0), 'Helvetica-Bold'))
                        
                        # Efeito Zebra sutil
                        for row_idx in range(1, len(tabela_tp_pdf)):
                            if row_idx % 2 == 0:
                                estilos_tabela.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#F8FAFC')))
                            else:
                                estilos_tabela.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.white))
                        
                        # Geração final
                        t_tp = Table(tabela_tp_pdf, colWidths=larguras_pdf_filtradas, repeatRows=1)
                        t_tp.setStyle(TableStyle(estilos_tabela))
                        
                        story_tp.append(t_tp)
                        doc_tp.build(story_tp)
                        
                        with open(caminho_temp_pdf_tp, "rb") as f_pdf_tp:
                            st.download_button("🔴 Baixar TEMPLATE FINAL em PDF (.pdf)", data=f_pdf_tp.read(), file_name=f"TEMPLATE_{mes_relatorio}_{ano_relatorio}.pdf", mime="application/pdf", use_container_width=True)

            # --- FLUXO DO BOTÃO 3: SERVIDORES ---
            elif st.session_state.relatorio_tipo_selecionado == "servidores":
                st.markdown("### 👤 Relatório Consolidado de **Servidores**")
                
                dados_completos_servidores = []
                conn = None
                try:
                    conn, cursor = conectar_banco()
                    
                    cursor.execute("""
                        SELECT DISTINCT
                            s.unidade_lotacao, 
                            s.nome_completo, 
                            s.cpf, 
                            e.numero_empenho,
                            e.valor_total as valor_total_empenho,
                            e.id as empenho_id,
                            s.id as servidor_id
                        FROM servidores s
                        JOIN diarias d ON d.servidor_id = s.id
                        JOIN empenhos e ON d.empenho_id = e.id
                        WHERE s.unidade_lotacao IS NOT NULL
                        
                        UNION
                        
                        SELECT DISTINCT
                            s.unidade_lotacao,
                            s.nome_completo,
                            s.cpf,
                            e.numero_empenho,
                            e.valor_total as valor_total_empenho,
                            e.id as empenho_id,
                            s.id as servidor_id
                        FROM empenhos e
                        JOIN servidores s ON s.unidade_lotacao IS NOT NULL
                        WHERE e.numero_empenho IS NOT NULL
                        ORDER BY unidade_lotacao ASC, nome_completo ASC
                    """)
                    dados_completos_servidores = cursor.fetchall()
                except Exception as e:
                    try:
                        cursor.execute("""
                            SELECT DISTINCT
                                s.unidade_lotacao, 
                                s.nome_completo, 
                                s.cpf, 
                                e.numero_empenho,
                                e.valor_total as valor_total_empenho,
                                e.id as empenho_id,
                                s.id as servidor_id
                            FROM servidores s
                            LEFT JOIN diarias d ON d.servidor_id = s.id
                            LEFT JOIN empenhos e ON d.empenho_id = e.id
                            WHERE s.unidade_lotacao IS NOT NULL
                            ORDER BY s.unidade_lotacao ASC, s.nome_completo ASC
                        """)
                        dados_completos_servidores = cursor.fetchall()
                    except Exception as err_fallback:
                        st.error(f"Erro ao buscar servidores e empenhos: {err_fallback}")
                finally:
                    if conn:
                        conn.close()
                
                if not dados_completos_servidores:
                    st.info("Nenhum servidor ou empenho localizado no sistema.")
                else:
                    df_servidores_full = pd.DataFrame(dados_completos_servidores, columns=[
                        "Unidade", "Policial", "CPF", "Empenho", "Valor_Total_Empenho", "Empenho_ID", "Servidor_ID"
                    ])
                    
                    df_servidores_full = df_servidores_full.drop_duplicates(subset=["Policial", "Empenho", "Unidade"])
                    
                    lista_unidades_totais = sorted(list(set(df_servidores_full["Unidade"].dropna().unique())))
                    
                    unidades_selecionadas = st.multiselect("Filtrar por Unidade de Lotação", options=lista_unidades_totais, default=lista_unidades_totais, key="filtro_unidades_servidores")
                    
                    if not unidades_selecionadas:
                        st.warning("⚠️ Selecione pelo menos uma unidade para gerar a visualização.")
                    else:
                        df_filtrado = df_servidores_full[df_servidores_full["Unidade"].isin(unidades_selecionadas)].copy()
                        
                        saldos_calculados = []
                        conn = None
                        try:
                            conn, cursor = conectar_banco()
                            for idx, row in df_filtrado.iterrows():
                                empenho_id = row["Empenho_ID"]
                                servidor_id = row["Servidor_ID"]
                                
                                if pd.isna(empenho_id) or empenho_id is None:
                                    saldos_calculados.append({
                                        "Policial": row["Policial"],
                                        "Empenho": "Sem Empenho Cadastrado",
                                        "Saldo": 0.0
                                    })
                                    continue
                                
                                cursor.execute("""
                                    SELECT SUM(valor_diaria)
                                    FROM diarias 
                                    WHERE empenho_id = ? AND servidor_id = ?
                                """, (int(empenho_id), int(servidor_id)))
                                soma_consumida = cursor.fetchone()[0]
                                soma_consumida = float(soma_consumida) if soma_consumida else 0.0
                                
                                valor_total_emp = float(row["Valor_Total_Empenho"]) if row["Valor_Total_Empenho"] else 0.0
                                saldo_atual = valor_total_emp - soma_consumida
                                
                                saldos_calculados.append({
                                    "Policial": row["Policial"],
                                    "Empenho": row["Empenho"] if row["Empenho"] else "Sem Empenho Cadastrado",
                                    "Saldo": max(0.0, saldo_atual)
                                })
                        except Exception as e:
                            st.error(f"Erro ao calcular saldos dos empenhos: {e}")
                        finally:
                            if conn:
                                conn.close()
                                
                        df_bruto_saldos = pd.DataFrame(saldos_calculados)
                        
                        if not df_bruto_saldos.empty:
                            df_consolidado_serv = df_bruto_saldos.groupby("Policial").agg({
                                "Saldo": "sum",
                                "Empenho": lambda x: ", ".join(sorted(list(set(x))))
                            }).reset_index()
                            
                            df_consolidado_serv["Qtd_DEAEVs"] = df_consolidado_serv["Saldo"].apply(lambda x: int(x // 360))
                            
                            st.markdown("#### Detalhamento de Saldos por Servidor")
                            st.dataframe(
                                df_consolidado_serv.style.format({"Saldo": "R$ {:,.2f}"}),
                                use_container_width=True,
                                hide_index=True
                            )
                            
                            st.markdown("---")
                            
                            caminho_temp_pdf_serv = os.path.join(DIRETORIO_ATUAL, "temp_servidores.pdf")
                            doc_serv = SimpleDocTemplate(caminho_temp_pdf_serv, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
                            story_serv = []
                            
                            estilos = getSampleStyleSheet()
                            estilo_titulo = ParagraphStyle('TitServ', parent=estilos['Heading3'], alignment=1, spaceAfter=5)
                            estilo_sub_unid = ParagraphStyle('SubUnid', parent=estilos['Normal'], alignment=1, fontSize=9, spaceAfter=20)
                            
                            story_serv.append(Paragraph(f"RELATÓRIO DE SALDOS DE SERVIDORES - {mes_relatorio}/{ano_relatorio}", estilo_titulo))
                            story_serv.append(Paragraph(f"Unidades Filtradas: {', '.join([u.split(' - ')[0] for u in unidades_selecionadas])}", estilo_sub_unid))
                            
                            tabela_dados_pdf = [["Nome do Servidor", "Empenho", "Qtd DEAEVs Restantes", "Total de DEAEVs"]]
                            
                            df_individual_pdf = df_bruto_saldos.copy().sort_values(by="Policial")
                            
                            estilos_tabela = [
                                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#729FCD')),
                                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                                ('ALIGN', (1,0), (-1,-1), 'CENTER'),
                                ('ALIGN', (2,0), (-1,-1), 'CENTER'),
                                ('ALIGN', (3,0), (-1,-1), 'CENTER'),
                                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                                ('FONTSIZE', (0,0), (-1,-1), 9),
                                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                                ('TOPPADDING', (0,0), (-1,-1), 6),
                                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                            ]
                            
                            indice_linha = 1
                            registro_agrupado = {}
                            somas_finais_servidor = {}
                            
                            for idx, row in df_individual_pdf.iterrows():
                                nome = row["Policial"][:45]
                                qtd_deae = int(row["Saldo"] // 360) if row["Saldo"] else 0
                                somas_finais_servidor[nome] = somas_finais_servidor.get(nome, 0) + qtd_deae
                            
                            for idx, row in df_individual_pdf.iterrows():
                                nome = row["Policial"][:45]
                                empenho = row["Empenho"]
                                qtd_deae = int(row["Saldo"] // 360) if row["Saldo"] else 0
                                total_deae = str(somas_finais_servidor[nome])
                                
                                tabela_dados_pdf.append([nome, empenho, str(qtd_deae), total_deae])
                                
                                if nome not in registro_agrupado:
                                    registro_agrupado[nome] = []
                                registro_agrupado[nome].append(indice_linha)
                                indice_linha += 1
                                
                            for nome, linhas in registro_agrupado.items():
                                if len(linhas) > 1:
                                    primeira_linha = linhas[0]
                                    ultima_linha = linhas[-1]
                                    
                                    estilos_tabela.append(('SPAN', (0, primeira_linha), (0, ultima_linha)))
                                    estilos_tabela.append(('SPAN', (3, primeira_linha), (3, ultima_linha)))
                                    
                                    estilos_tabela.append(('VALIGN', (0, primeira_linha), (0, ultima_linha), 'MIDDLE'))
                                    estilos_tabela.append(('VALIGN', (3, primeira_linha), (3, ultima_linha), 'MIDDLE'))
                                    
                                    estilos_tabela.append(('FONTNAME', (3, primeira_linha), (3, ultima_linha), 'Helvetica-Bold'))
                            
                            t_serv = Table(tabela_dados_pdf, colWidths=[200, 130, 125, 125], repeatRows=1)
                            t_serv.setStyle(TableStyle(estilos_tabela))
                            story_serv.append(t_serv)
                            doc_serv.build(story_serv)
                            
                            with open(caminho_temp_pdf_serv, "rb") as f_pdf_serv:
                                st.download_button(
                                    "🔴 Exportar Relatório de Servidores para PDF (.pdf)", 
                                    data=f_pdf_serv.read(), 
                                    file_name=f"relatorio_servidores_{mes_relatorio}_{ano_relatorio}.pdf", 
                                    mime="application/pdf", 
                                    use_container_width=True
                                )
                        
            # --- FLUXO DO BOTÃO 4: EMPENHOS UTILIZADOS ---
            elif st.session_state.relatorio_tipo_selecionado == "empenhos_utilizados":
                st.markdown("### 🔍 Opções para: **EMPENHOS UTILIZADOS**")
                st.caption(f"Empenhos que tiveram ao menos uma diária (DEAEV) lançada em **{mes_relatorio}/{ano_relatorio}** (filtro definido no topo da página). Use o botão abaixo para baixar todos os PDFs desses empenhos em um único pacote, prontos para a prestação de contas.")

                conn_eu = None
                try:
                    conn_eu, cursor_eu = conectar_banco()
                    cursor_eu.execute("""
                        SELECT 
                            e.id,
                            e.numero_empenho,
                            e.valor_total,
                            e.valor_disponivel,
                            e.data_empenho,
                            e.caminho_pdf,
                            GROUP_CONCAT(DISTINCT s.nome_completo) AS servidores_beneficiados,
                            SUM(d.valor_diaria) AS valor_utilizado_mes
                        FROM empenhos e
                        JOIN diarias d ON d.empenho_id = e.id
                        JOIN servidores s ON d.servidor_id = s.id
                        WHERE strftime('%Y-%m', d.data_diaria) = ?
                        GROUP BY e.id
                        ORDER BY e.numero_empenho ASC
                    """, (competencia_busca,))
                    empenhos_utilizados_mes = cursor_eu.fetchall()
                except Exception as e:
                    st.error(f"Erro ao buscar os empenhos utilizados no período: {e}")
                    empenhos_utilizados_mes = []
                finally:
                    if conn_eu:
                        conn_eu.close()

                if not empenhos_utilizados_mes:
                    st.info(f"Nenhum empenho foi utilizado (não há diárias lançadas) em {mes_relatorio}/{ano_relatorio}.")
                else:
                    df_empenhos_utilizados = pd.DataFrame(
                        empenhos_utilizados_mes,
                        columns=["ID", "Empenho", "Valor Total", "Valor Disponível Atual", "Data do Empenho", "Caminho PDF", "Servidores Beneficiados", "Valor Utilizado no Mês"]
                    )

                    st.markdown(f"#### 📄 {len(df_empenhos_utilizados)} empenho(s) utilizado(s) em {mes_relatorio}/{ano_relatorio}")
                    st.dataframe(
                        df_empenhos_utilizados[["Empenho", "Servidores Beneficiados", "Valor Utilizado no Mês", "Valor Total", "Valor Disponível Atual"]].style.format({
                            "Valor Utilizado no Mês": "R$ {:,.2f}",
                            "Valor Total": "R$ {:,.2f}",
                            "Valor Disponível Atual": "R$ {:,.2f}"
                        }),
                        use_container_width=True,
                        hide_index=True
                    )

                    st.markdown("---")

                    # Monta o pacote ZIP em memória com os PDFs dos empenhos utilizados no período
                    buffer_zip = io.BytesIO()
                    empenhos_sem_pdf = []
                    nomes_usados_zip = set()

                    with zipfile.ZipFile(buffer_zip, "w", zipfile.ZIP_DEFLATED) as pacote_zip:
                        for _, linha_emp in df_empenhos_utilizados.iterrows():
                            caminho_pdf_emp = linha_emp["Caminho PDF"]
                            numero_emp = str(linha_emp["Empenho"])

                            if caminho_pdf_emp and os.path.exists(caminho_pdf_emp):
                                nome_base_zip = f"{numero_emp}.pdf".replace("/", "-")
                                nome_final_zip = nome_base_zip
                                contador_dup = 1
                                while nome_final_zip in nomes_usados_zip:
                                    contador_dup += 1
                                    nome_final_zip = f"{numero_emp}_{contador_dup}.pdf".replace("/", "-")
                                nomes_usados_zip.add(nome_final_zip)
                                pacote_zip.write(caminho_pdf_emp, arcname=nome_final_zip)
                            else:
                                empenhos_sem_pdf.append(numero_emp)

                        # Inclui uma planilha-resumo dentro do próprio pacote, útil para conferência na prestação de contas
                        buffer_resumo = io.BytesIO()
                        wb_resumo = Workbook()
                        ws_resumo = wb_resumo.active
                        ws_resumo.title = "Resumo"
                        ws_resumo.append(["Empenho", "Servidores Beneficiados", "Valor Utilizado no Mês", "Valor Total do Empenho", "Valor Disponível Atual", "Data do Empenho"])
                        for col_idx in range(1, 7):
                            celula_cab = ws_resumo.cell(row=1, column=col_idx)
                            celula_cab.font = Font(bold=True, color="FFFFFF")
                            celula_cab.fill = PatternFill(start_color="729FCD", end_color="729FCD", fill_type="solid")
                        for _, linha_emp in df_empenhos_utilizados.iterrows():
                            ws_resumo.append([
                                linha_emp["Empenho"],
                                linha_emp["Servidores Beneficiados"],
                                linha_emp["Valor Utilizado no Mês"],
                                linha_emp["Valor Total"],
                                linha_emp["Valor Disponível Atual"],
                                linha_emp["Data do Empenho"]
                            ])
                        for coluna_cel in ws_resumo.columns:
                            comprimento_max = max(len(str(c.value)) if c.value is not None else 0 for c in coluna_cel)
                            ws_resumo.column_dimensions[get_column_letter(coluna_cel[0].column)].width = comprimento_max + 4
                        wb_resumo.save(buffer_resumo)
                        pacote_zip.writestr(f"resumo_empenhos_utilizados_{mes_relatorio}_{ano_relatorio}.xlsx", buffer_resumo.getvalue())

                    if empenhos_sem_pdf:
                        st.warning(f"⚠️ Os seguintes empenhos não possuem arquivo PDF anexado no sistema e não entraram no pacote: {', '.join(empenhos_sem_pdf)}")

                    st.download_button(
                        "📦 Baixar Pacote de Empenhos Utilizados (.zip)",
                        data=buffer_zip.getvalue(),
                        file_name=f"empenhos_utilizados_{mes_relatorio}_{ano_relatorio}.zip",
                        mime="application/zip",
                        use_container_width=True,
                        type="primary"
                    )

    # =====================================================================
    # Bloco Sistema
    # =====================================================================
    elif st.session_state.menu_ativo == "sistema":
        st.markdown("<h2>⚙️ Parâmetros do Sistema</h2>", unsafe_allow_html=True)
        tab_unidades, tab_bancos = st.tabs(["🏛️ Unidades de Lotação", "🏦 Bancos Cadastrados"])
        
        with tab_unidades:
            st.markdown("#### Gerenciar Unidades Operacionais / Administrativas")
            with st.form("form_nova_unidade", clear_on_submit=True):
                nova_unidade_txt = st.text_input("Nome da Nova Unidade")
                if st.form_submit_button("➕ Adicionar Unidade"):
                    if not nova_unidade_txt.strip(): st.error("Digite um nome válido.")
                    else:
                        conn, cursor = conectar_banco()
                        try:
                            cursor.execute("INSERT INTO unidades (nome_unidade) VALUES (?)", (nova_unidade_txt.strip().upper(),))
                            conn.commit()
                            st.success("Unidade adicionada com sucesso!")
                        except sqlite3.IntegrityError: st.error("Esta unidade já está cadastrada.")
                        finally: conn.close(); st.rerun()
            
            conn, cursor = conectar_banco()
            cursor.execute("SELECT id, nome_unidade FROM unidades ORDER BY nome_unidade ASC")
            unidades_banco = cursor.fetchall()
            conn.close()
            if unidades_banco:
                st.dataframe(pd.DataFrame(unidades_banco, columns=["ID", "Nome da Unidade"]), use_container_width=True, hide_index=True)
                
        with tab_bancos:
            st.markdown("#### Gerenciar Instituições Bancárias")
            with st.form("form_novo_banco", clear_on_submit=True):
                col_c, col_n = st.columns([3, 9])
                with col_c: cod_banco_txt = st.text_input("Código Compe (3 dígitos)", max_chars=3)
                with col_n: nome_banco_txt = st.text_input("Nome Comercial do Banco")
                if st.form_submit_button("➕ Adicionar Banco"):
                    if not cod_banco_txt.strip() or not nome_banco_txt.strip(): st.error("Ambos os campos são obrigatórios.")
                    else:
                        conn, cursor = conectar_banco()
                        try:
                            cursor.execute("INSERT INTO bancos (codigo, nome_banco) VALUES (?, ?)", (cod_banco_txt.strip(), nome_banco_txt.strip().upper()))
                            conn.commit()
                            st.success("Banco cadastrado!")
                        except sqlite3.IntegrityError: st.error("Este código de banco já existe.")
                        finally: conn.close(); st.rerun()
            
            conn, cursor = conectar_banco()
            cursor.execute("SELECT codigo, nome_banco FROM bancos ORDER BY codigo ASC")
            bancos_banco = cursor.fetchall()
            conn.close()
            if bancos_banco:
                st.dataframe(pd.DataFrame(bancos_banco, columns=["Código Compe", "Nome do Banco"]), use_container_width=True, hide_index=True)